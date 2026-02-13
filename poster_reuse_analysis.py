#!/usr/bin/env python3
"""
Poster Reuse Analysis Script
============================
Comprehensive statistical analysis of poster metadata and engagement metrics.

This script performs:
1. Citation data cleaning (removing false positives)
2. Descriptive statistics for views, downloads, citations
3. Bivariate analysis (Spearman's, Mann-Whitney U, Point-Biserial, Chi-Square)
4. Multivariate analysis (Negative Binomial, Zero-Inflated Negative Binomial)
5. Publication-ready figures and tables

Author: James O'Neill
Date: 2026-02-04
"""

import pandas as pd
import numpy as np
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Optional

# Statistical libraries
from scipy import stats
from scipy.stats import spearmanr, mannwhitneyu, pointbiserialr, chi2_contingency
import statsmodels.api as sm
from statsmodels.discrete.count_model import ZeroInflatedNegativeBinomialP
from statsmodels.discrete.discrete_model import NegativeBinomial

# Visualization
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import seaborn as sns

# Suppress warnings for cleaner output
import warnings
warnings.filterwarnings('ignore')

# =============================================================================
# CONFIGURATION
# =============================================================================

# Data paths - using the Google Drive dataset
BASE_PATH = Path("/home/joneill/Nextcloud/vaults/jmind/calmi2/poster_science/poster-sharing-reuse-paper")
DATASET_PATH = BASE_PATH / "poster_metadata" / "Paper - poster metadata quality" / "dataset"
OUTPUT_PATH = BASE_PATH / "poster-sharing-reuse-paper-code" / "outputs" / "analysis_results"

# Input files
COMBINED_TABLE = DATASET_PATH / "derivative" / "poster-reuse" / "combined-table-for-reuse-analysis.csv"
CITATIONS_NDJSON = DATASET_PATH / "primary" / "citations" / "posters-citations.ndjson"
FIGSHARE_CSV = DATASET_PATH / "derivative" / "poster-metadata" / "figshare.csv"
ZENODO_CSV = DATASET_PATH / "derivative" / "poster-metadata" / "zenodo.csv"

# Metadata columns for analysis (from the paper's Methods section)
METADATA_COLS_CONTINUOUS = [
    "authors_affiliation_percentage",
    "authors_orcid_percentage",
    "description_words_count",
    "description_characters_count",
    "keywords_freetext_count",
    "keywords_controlled_vocabularies_count",
    "total_keywords_count",
    "references_no_identifiers_count",
    "references_with_identifiers_count",
    "total_references_count"
]

METADATA_COLS_BOOLEAN = [
    "has_title",
    "has_description",
    "has_license",
    "has_funding_info",
    "has_conference_acronym",
    "has_conference_dates",
    "has_conference_place",
    "has_conference_session",
    "has_conference_session_part",
    "has_conference_title",
    "has_conference_website"
]

OUTCOME_COLS = ["views", "downloads", "citations"]

# =============================================================================
# LOGGING UTILITIES
# =============================================================================

def log(message: str, level: str = "INFO") -> None:
    """Print timestamped log message."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] [{level}] {message}")

def log_section(title: str) -> None:
    """Print section header."""
    print("\n" + "=" * 80)
    log(title, "SECTION")
    print("=" * 80)

# =============================================================================
# DATA LOADING AND CLEANING
# =============================================================================

def load_combined_data() -> pd.DataFrame:
    """Load the combined poster metadata table."""
    log(f"Loading combined data from {COMBINED_TABLE}")
    df = pd.read_csv(COMBINED_TABLE, low_memory=False)
    log(f"Loaded {len(df):,} posters with {len(df.columns)} columns")
    return df

def load_citations_ndjson() -> pd.DataFrame:
    """Load citations from NDJSON file."""
    log(f"Loading citations from {CITATIONS_NDJSON}")
    records = []
    with open(CITATIONS_NDJSON, 'r', encoding='utf-8') as f:
        for line in f:
            records.append(json.loads(line.strip()))
    df = pd.DataFrame(records)
    log(f"Loaded {len(df):,} citation records")
    return df

def clean_citations(df_citations: pd.DataFrame, df_posters: pd.DataFrame) -> Tuple[pd.DataFrame, Dict]:
    """
    Clean citations by removing false positives where citation_year < poster creation_year.
    
    Returns:
        - Cleaned citations DataFrame
        - Dictionary with cleaning statistics
    """
    log("Cleaning citations data...")
    
    # Create poster DOI to creation_year mapping
    poster_years = df_posters.set_index('doi_norm')['creation_year'].to_dict()
    
    # Add poster creation year to citations
    df_citations = df_citations.copy()
    df_citations['poster_creation_year'] = df_citations['poster_doi'].map(
        lambda x: poster_years.get(x.lower() if isinstance(x, str) else x)
    )
    
    # Convert citation_year to numeric
    df_citations['citation_year'] = pd.to_numeric(df_citations['citation_year'], errors='coerce')
    
    # Count before cleaning
    total_before = len(df_citations)
    
    # Identify false positives (citation year < poster creation year)
    false_positives_mask = (
        df_citations['citation_year'].notna() & 
        df_citations['poster_creation_year'].notna() & 
        (df_citations['citation_year'] < df_citations['poster_creation_year'])
    )
    
    n_false_positives = false_positives_mask.sum()
    
    # Log examples of false positives
    if n_false_positives > 0:
        log(f"Found {n_false_positives} false positive citations (citation_year < creation_year)")
        examples = df_citations[false_positives_mask][['poster_doi', 'citation_year', 'poster_creation_year']].head(5)
        log("Examples of false positives:")
        for _, row in examples.iterrows():
            log(f"  DOI: {row['poster_doi'][:50]}... | Citation: {row['citation_year']} < Poster: {row['poster_creation_year']}")
    
    # Remove false positives
    df_clean = df_citations[~false_positives_mask].copy()
    
    stats = {
        'total_before': total_before,
        'false_positives_removed': n_false_positives,
        'total_after': len(df_clean),
        'pct_removed': (n_false_positives / total_before * 100) if total_before > 0 else 0
    }
    
    log(f"Cleaned citations: {stats['total_after']:,} remain ({stats['pct_removed']:.2f}% removed)")
    
    return df_clean, stats

def save_clean_citations(df_clean: pd.DataFrame, output_path: Path) -> None:
    """Save cleaned citations to NDJSON."""
    output_file = output_path / "posters-citations-clean.ndjson"
    log(f"Saving cleaned citations to {output_file}")
    
    with open(output_file, 'w', encoding='utf-8') as f:
        for _, row in df_clean.iterrows():
            record = row.dropna().to_dict()
            # Remove the helper column
            record.pop('poster_creation_year', None)
            f.write(json.dumps(record, ensure_ascii=False) + '\n')
    
    log(f"Saved {len(df_clean):,} cleaned citation records")

def update_citation_counts(df_posters: pd.DataFrame, df_citations_clean: pd.DataFrame) -> pd.DataFrame:
    """Update citation counts in poster table using cleaned citations."""
    log("Updating citation counts with cleaned data...")
    
    # Count citations per poster DOI
    citation_counts = df_citations_clean.groupby('poster_doi').size()
    
    # Map to posters
    df_posters = df_posters.copy()
    df_posters['citations_clean'] = df_posters['doi_norm'].map(citation_counts).fillna(0).astype(int)
    
    # Compare with original
    original_sum = df_posters['citations'].sum() if 'citations' in df_posters.columns else 0
    clean_sum = df_posters['citations_clean'].sum()
    
    log(f"Citation counts: Original={original_sum:,}, Cleaned={clean_sum:,}, Diff={original_sum - clean_sum:,}")
    
    return df_posters

# =============================================================================
# DESCRIPTIVE STATISTICS
# =============================================================================

def compute_distribution_stats(series: pd.Series, name: str) -> Dict:
    """Compute comprehensive distribution statistics for a series."""
    s = series.dropna()
    
    stats_dict = {
        'variable': name,
        'n': int(s.count()),
        'sum': float(s.sum()),
        'mean': float(s.mean()),
        'median': float(s.median()),
        'std': float(s.std(ddof=1)),
        'min': float(s.min()),
        'max': float(s.max()),
        'p25': float(s.quantile(0.25)),
        'p75': float(s.quantile(0.75)),
        'p90': float(s.quantile(0.90)),
        'p95': float(s.quantile(0.95)),
        'p99': float(s.quantile(0.99)),
    }
    
    # Zero inflation analysis
    n_zero = int((s == 0).sum())
    stats_dict['n_zero'] = n_zero
    stats_dict['pct_zero'] = 100.0 * n_zero / max(int(s.count()), 1)
    
    # Non-zero subset stats
    s_nz = s[s > 0]
    stats_dict['n_nonzero'] = int(s_nz.count())
    stats_dict['pct_nonzero'] = 100.0 * int(s_nz.count()) / max(int(s.count()), 1)
    
    if len(s_nz) > 0:
        stats_dict['mean_nonzero'] = float(s_nz.mean())
        stats_dict['median_nonzero'] = float(s_nz.median())
    else:
        stats_dict['mean_nonzero'] = np.nan
        stats_dict['median_nonzero'] = np.nan
    
    return stats_dict

def compute_concentration(series: pd.Series, top_pcts: List[float] = [0.01, 0.05, 0.10]) -> Dict:
    """Compute concentration metrics (what % of total is held by top X%)."""
    s = series.dropna().clip(lower=0)
    total = s.sum()
    
    if total <= 0:
        return {f'top_{int(p*100)}pct_share': np.nan for p in top_pcts}
    
    s_sorted = s.sort_values(ascending=False)
    n = len(s_sorted)
    
    concentration = {}
    for pct in top_pcts:
        k = max(int(np.ceil(pct * n)), 1)
        top_sum = s_sorted.head(k).sum()
        concentration[f'top_{int(pct*100)}pct_share'] = 100.0 * top_sum / total
    
    return concentration

def generate_table2(df: pd.DataFrame) -> pd.DataFrame:
    """Generate Table 2: Distribution statistics for views, downloads, citations."""
    log("Generating Table 2: Distribution of engagement metrics...")
    
    results = []
    for col in OUTCOME_COLS:
        if col in df.columns:
            stats = compute_distribution_stats(df[col], col)
            concentration = compute_concentration(df[col])
            stats.update(concentration)
            results.append(stats)
    
    table2 = pd.DataFrame(results)
    log(f"Table 2 generated with {len(table2)} metrics")
    return table2

# =============================================================================
# BIVARIATE ANALYSIS: VIEWS & DOWNLOADS
# =============================================================================

def bivariate_continuous_spearman(df: pd.DataFrame, outcome: str, predictors: List[str]) -> pd.DataFrame:
    """
    Spearman's Rank Correlation for continuous metadata vs outcome.
    Used for views/downloads analysis.
    """
    results = []
    for pred in predictors:
        if pred not in df.columns or outcome not in df.columns:
            continue
        
        mask = df[pred].notna() & df[outcome].notna()
        x = df.loc[mask, pred]
        y = df.loc[mask, outcome]
        
        if len(x) < 3:
            continue
        
        rho, pvalue = spearmanr(x, y)
        
        results.append({
            'predictor': pred,
            'outcome': outcome,
            'test': 'Spearman',
            'statistic': rho,
            'p_value': pvalue,
            'n': len(x),
            'significant_05': pvalue < 0.05,
            'significant_01': pvalue < 0.01,
            'significant_001': pvalue < 0.001
        })
    
    return pd.DataFrame(results)

def bivariate_boolean_mannwhitney(df: pd.DataFrame, outcome: str, predictors: List[str]) -> pd.DataFrame:
    """
    Mann-Whitney U Test for boolean metadata vs outcome.
    Used for views/downloads analysis.
    """
    results = []
    for pred in predictors:
        if pred not in df.columns or outcome not in df.columns:
            continue
        
        mask = df[pred].notna() & df[outcome].notna()
        group_true = df.loc[mask & (df[pred] == True), outcome]
        group_false = df.loc[mask & (df[pred] == False), outcome]
        
        if len(group_true) < 2 or len(group_false) < 2:
            continue
        
        statistic, pvalue = mannwhitneyu(group_true, group_false, alternative='two-sided')
        
        # Effect size: rank-biserial correlation
        n1, n2 = len(group_true), len(group_false)
        r = 1 - (2 * statistic) / (n1 * n2)
        
        results.append({
            'predictor': pred,
            'outcome': outcome,
            'test': 'Mann-Whitney U',
            'statistic': statistic,
            'p_value': pvalue,
            'effect_size_r': r,
            'n_true': n1,
            'n_false': n2,
            'median_true': group_true.median(),
            'median_false': group_false.median(),
            'significant_05': pvalue < 0.05,
            'significant_01': pvalue < 0.01,
            'significant_001': pvalue < 0.001
        })
    
    return pd.DataFrame(results)

# =============================================================================
# BIVARIATE ANALYSIS: CITATIONS (BINARY - WAS CITED?)
# =============================================================================

def bivariate_binary_pointbiserial(df: pd.DataFrame, predictors: List[str]) -> pd.DataFrame:
    """
    Point-Biserial Correlation for continuous metadata vs binary "was cited".
    """
    df = df.copy()
    df['was_cited'] = (df['citations'] > 0).astype(int)
    
    results = []
    for pred in predictors:
        if pred not in df.columns:
            continue
        
        mask = df[pred].notna() & df['was_cited'].notna()
        x = df.loc[mask, pred]
        y = df.loc[mask, 'was_cited']
        
        if len(x) < 3:
            continue
        
        r, pvalue = pointbiserialr(y, x)
        
        results.append({
            'predictor': pred,
            'outcome': 'was_cited',
            'test': 'Point-Biserial',
            'correlation': r,
            'p_value': pvalue,
            'n': len(x),
            'significant_05': pvalue < 0.05,
            'significant_01': pvalue < 0.01,
            'significant_001': pvalue < 0.001
        })
    
    return pd.DataFrame(results)

def bivariate_binary_chisquare(df: pd.DataFrame, predictors: List[str]) -> pd.DataFrame:
    """
    Chi-Square Test of Independence for boolean metadata vs binary "was cited".
    """
    df = df.copy()
    df['was_cited'] = (df['citations'] > 0).astype(int)
    
    results = []
    for pred in predictors:
        if pred not in df.columns:
            continue
        
        mask = df[pred].notna() & df['was_cited'].notna()
        subset = df.loc[mask, [pred, 'was_cited']]
        
        if len(subset) < 10:
            continue
        
        # Create contingency table
        contingency = pd.crosstab(subset[pred], subset['was_cited'])
        
        if contingency.shape[0] < 2 or contingency.shape[1] < 2:
            continue
        
        chi2, pvalue, dof, expected = chi2_contingency(contingency)
        
        # Cramér's V effect size
        n = contingency.sum().sum()
        min_dim = min(contingency.shape[0] - 1, contingency.shape[1] - 1)
        cramers_v = np.sqrt(chi2 / (n * min_dim)) if min_dim > 0 else np.nan
        
        results.append({
            'predictor': pred,
            'outcome': 'was_cited',
            'test': 'Chi-Square',
            'chi2': chi2,
            'p_value': pvalue,
            'dof': dof,
            'cramers_v': cramers_v,
            'n': n,
            'significant_05': pvalue < 0.05,
            'significant_01': pvalue < 0.01,
            'significant_001': pvalue < 0.001
        })
    
    return pd.DataFrame(results)

# =============================================================================
# BIVARIATE ANALYSIS: CITATIONS (COUNT - AMONG CITED POSTERS)
# =============================================================================

def bivariate_citations_count(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Analyze citation counts among posters that have at least one citation.
    Uses Spearman for continuous and Mann-Whitney U for boolean.
    """
    # Filter to cited posters only
    df_cited = df[df['citations'] > 0].copy()
    log(f"Analyzing {len(df_cited):,} cited posters (out of {len(df):,} total)")
    
    # Spearman for continuous
    spearman_results = bivariate_continuous_spearman(df_cited, 'citations', METADATA_COLS_CONTINUOUS)
    
    # Mann-Whitney U for boolean
    mannwhitney_results = bivariate_boolean_mannwhitney(df_cited, 'citations', METADATA_COLS_BOOLEAN)
    
    return spearman_results, mannwhitney_results

# =============================================================================
# MULTIVARIATE ANALYSIS: NEGATIVE BINOMIAL REGRESSION (ROBUST VERSION)
# =============================================================================

from sklearn.preprocessing import StandardScaler

def estimate_dispersion(y: pd.Series) -> float:
    """
    Estimate NB dispersion parameter from intercept-only model.
    Uses method of moments: alpha = (var - mean) / mean^2
    """
    mean_y = y.mean()
    var_y = y.var()
    if mean_y > 0:
        alpha = max((var_y - mean_y) / (mean_y ** 2), 0.01)
    else:
        alpha = 1.0
    return alpha

def get_clean_predictors(predictors: List[str]) -> List[str]:
    """
    Remove quasi-invariant and redundant predictors to reduce collinearity.
    """
    # Remove problematic predictors
    exclude = [
        'has_title',  # Nearly 100% True - causes separation
        'description_characters_count',  # Redundant with description_words_count
    ]
    return [p for p in predictors if p not in exclude]

def negative_binomial_regression(df: pd.DataFrame, outcome: str, predictors: List[str]) -> Dict:
    """
    Fit Negative Binomial regression with fixed dispersion parameter.
    
    This approach:
    1. Estimates α (dispersion) from an intercept-only model first
    2. Fixes α when fitting the full model with all predictors
    3. Always converges (no iterative α estimation in full model)
    4. Gives consistent, efficient coefficient estimates
    
    Reference: Cameron & Trivedi (2013), Regression Analysis of Count Data
    
    Why fixed α instead of estimated?
    - When overdispersion is extreme (var/mean > 1000), MLE of α becomes unstable
    - Fixing α from a simpler model ensures convergence
    - β estimates remain consistent regardless of α specification
    """
    log(f"Fitting Negative Binomial Regression with fixed α for {outcome}...")
    
    # Clean predictors
    clean_predictors = get_clean_predictors([p for p in predictors if p in df.columns])
    
    # Prepare data
    df_model = df[clean_predictors + [outcome]].dropna().copy()
    
    # Convert boolean to int
    for col in df_model.columns:
        if df_model[col].dtype == 'bool':
            df_model[col] = df_model[col].astype(int)
    
    y = df_model[outcome].astype(float)
    X = df_model[clean_predictors].astype(float)
    
    # Standardize continuous predictors
    continuous_cols = [c for c in clean_predictors if not c.startswith('has_')]
    scaler = StandardScaler()
    X_scaled = X.copy()
    X_scaled[continuous_cols] = scaler.fit_transform(X[continuous_cols])
    X_scaled = sm.add_constant(X_scaled)
    
    log(f"Model data: {len(df_model):,} observations, {len(clean_predictors)} predictors")
    
    # Calculate overdispersion for reporting
    mean_y = y.mean()
    var_y = y.var()
    dispersion_ratio = var_y / mean_y if mean_y > 0 else np.nan
    log(f"Overdispersion ratio (var/mean): {dispersion_ratio:.2f}")
    
    # Step 1: Estimate α from intercept-only model using method of moments
    # α = (var - mean) / mean² for NB2 parameterization
    if mean_y > 0:
        alpha_estimate = max((var_y - mean_y) / (mean_y ** 2), 0.1)
        # Cap at reasonable value to avoid numerical issues
        alpha_estimate = min(alpha_estimate, 10.0)
    else:
        alpha_estimate = 1.0
    
    log(f"Estimated dispersion α (fixed): {alpha_estimate:.4f}")
    
    try:
        # Fit Negative Binomial GLM with fixed α
        # In statsmodels, NegativeBinomial family takes alpha as the dispersion parameter
        nb_family = sm.families.NegativeBinomial(alpha=alpha_estimate)
        nb_model = sm.GLM(y, X_scaled, family=nb_family)
        nb_fit = nb_model.fit(maxiter=200)
        
        # Extract results
        results = {
            'outcome': outcome,
            'model_type': 'Negative Binomial with fixed dispersion',
            'n_obs': int(len(y)),
            'alpha_fixed': float(alpha_estimate),
            'overdispersion_ratio': float(dispersion_ratio),
            'deviance': float(nb_fit.deviance),
            'pearson_chi2': float(nb_fit.pearson_chi2),
            'aic': float(nb_fit.aic) if hasattr(nb_fit, 'aic') else np.nan,
            'converged': nb_fit.converged,
            'predictors_used': clean_predictors,
            'predictors_excluded': [p for p in predictors if p not in clean_predictors],
            'scaling_applied': True,
            'coefficients': {}
        }
        
        # Calculate IRR with standard errors
        for param in nb_fit.params.index:
            coef = nb_fit.params[param]
            se = nb_fit.bse[param]
            z_val = coef / se if se > 0 else np.nan
            pval = nb_fit.pvalues[param] if param in nb_fit.pvalues else 2 * (1 - stats.norm.cdf(abs(z_val)))
            
            irr = np.exp(coef)
            irr_lower = np.exp(coef - 1.96 * se)
            irr_upper = np.exp(coef + 1.96 * se)
            
            results['coefficients'][param] = {
                'coefficient': float(coef),
                'std_error': float(se),
                'z_value': float(z_val),
                'irr': float(irr),
                'irr_ci_lower': float(irr_lower),
                'irr_ci_upper': float(irr_upper),
                'p_value': float(pval),
                'significant_05': bool(pval < 0.05),
                'significant_01': bool(pval < 0.01),
                'significant_001': bool(pval < 0.001)
            }
        
        log(f"Model converged: {results['converged']}")
        return results
        
    except Exception as e:
        log(f"Error fitting model: {e}", "ERROR")
        return {'outcome': outcome, 'error': str(e)}

# =============================================================================
# MULTIVARIATE ANALYSIS: TWO-PART HURDLE MODEL FOR CITATIONS
# =============================================================================

def hurdle_model_citations(df: pd.DataFrame, predictors: List[str]) -> Dict:
    """
    Fit a two-part hurdle model for citations:
    
    Part 1: Logistic regression for P(cited | X)
            - Models whether a poster receives ANY citation
            
    Part 2: Truncated Negative Binomial for E[citations | cited, X]
            - Models citation COUNT among cited posters only
            - Uses NB with fixed α (estimated from intercept-only model)
            
    This approach is more interpretable than ZINB and converges reliably.
    Reference: Cameron & Trivedi (2013), Regression Analysis of Count Data
    """
    log("Fitting Two-Part Hurdle Model for citations...")
    
    # Clean predictors
    clean_predictors = get_clean_predictors([p for p in predictors if p in df.columns])
    
    # Prepare data
    df_model = df[clean_predictors + ['citations']].dropna().copy()
    
    # Convert boolean to int
    for col in df_model.columns:
        if df_model[col].dtype == 'bool':
            df_model[col] = df_model[col].astype(int)
    
    # Create binary outcome
    df_model['was_cited'] = (df_model['citations'] > 0).astype(int)
    
    log(f"Total posters: {len(df_model):,}")
    log(f"Cited posters: {df_model['was_cited'].sum():,} ({df_model['was_cited'].mean()*100:.1f}%)")
    
    # Standardize continuous predictors
    continuous_cols = [c for c in clean_predictors if not c.startswith('has_')]
    scaler = StandardScaler()
    X = df_model[clean_predictors].astype(float)
    X_scaled = X.copy()
    X_scaled[continuous_cols] = scaler.fit_transform(X[continuous_cols])
    X_scaled = sm.add_constant(X_scaled)
    
    results = {
        'outcome': 'citations',
        'n_total': int(len(df_model)),
        'n_cited': int(df_model['was_cited'].sum()),
        'pct_cited': float(df_model['was_cited'].mean() * 100),
        'predictors_used': clean_predictors,
        'predictors_excluded': [p for p in predictors if p not in clean_predictors],
        'scaling_applied': True,
        'part1_logistic': {},
        'part2_count': {}
    }
    
    # =========================================================================
    # PART 1: Logistic Regression for P(cited)
    # =========================================================================
    log("Part 1: Fitting Logistic Regression for citation probability...")
    
    try:
        y_binary = df_model['was_cited'].astype(float)
        logit_model = sm.GLM(y_binary, X_scaled, family=sm.families.Binomial())
        logit_fit = logit_model.fit(maxiter=200)
        
        results['part1_logistic'] = {
            'n_obs': int(len(y_binary)),
            'converged': logit_fit.converged,
            'deviance': float(logit_fit.deviance),
            'aic': float(logit_fit.aic),
            'pseudo_r2': float(1 - logit_fit.deviance / logit_fit.null_deviance),
            'coefficients': {}
        }
        
        for param in logit_fit.params.index:
            coef = logit_fit.params[param]
            se = logit_fit.bse[param]
            pval = logit_fit.pvalues[param]
            
            odds_ratio = np.exp(coef)
            or_lower = np.exp(coef - 1.96 * se)
            or_upper = np.exp(coef + 1.96 * se)
            
            results['part1_logistic']['coefficients'][param] = {
                'coefficient': float(coef),
                'std_error': float(se),
                'z_value': float(coef / se) if se > 0 else np.nan,
                'odds_ratio': float(odds_ratio),
                'or_ci_lower': float(or_lower),
                'or_ci_upper': float(or_upper),
                'p_value': float(pval),
                'significant_05': bool(pval < 0.05),
                'significant_01': bool(pval < 0.01),
                'significant_001': bool(pval < 0.001)
            }
        
        log(f"Part 1 converged: {logit_fit.converged}, Pseudo-R²: {results['part1_logistic']['pseudo_r2']:.4f}")
        
    except Exception as e:
        log(f"Error in Part 1 (Logistic): {e}", "ERROR")
        results['part1_logistic']['error'] = str(e)
    
    # =========================================================================
    # PART 2: Negative Binomial with fixed α for citation count among cited posters
    # =========================================================================
    log("Part 2: Fitting Negative Binomial with fixed α for citation count (cited posters only)...")
    
    try:
        # Filter to cited posters only
        df_cited = df_model[df_model['was_cited'] == 1].copy()
        y_count = df_cited['citations'].astype(float)
        
        X_cited = df_cited[clean_predictors].astype(float)
        X_cited_scaled = X_cited.copy()
        X_cited_scaled[continuous_cols] = scaler.transform(X_cited[continuous_cols])
        X_cited_scaled = sm.add_constant(X_cited_scaled)
        
        # Calculate overdispersion for reporting
        mean_y = y_count.mean()
        var_y = y_count.var()
        dispersion_ratio = var_y / mean_y if mean_y > 0 else np.nan
        log(f"Part 2 overdispersion ratio: {dispersion_ratio:.2f}")
        
        # Estimate α from method of moments
        if mean_y > 0:
            alpha_estimate = max((var_y - mean_y) / (mean_y ** 2), 0.1)
            alpha_estimate = min(alpha_estimate, 10.0)
        else:
            alpha_estimate = 1.0
        
        log(f"Part 2 estimated dispersion α (fixed): {alpha_estimate:.4f}")
        
        # Fit Negative Binomial GLM with fixed α
        nb_family = sm.families.NegativeBinomial(alpha=alpha_estimate)
        nb_model = sm.GLM(y_count, X_cited_scaled, family=nb_family)
        nb_fit = nb_model.fit(maxiter=200)
        
        results['part2_count'] = {
            'n_obs': int(len(y_count)),
            'model_type': 'Negative Binomial with fixed dispersion',
            'alpha_fixed': float(alpha_estimate),
            'overdispersion_ratio': float(dispersion_ratio),
            'converged': nb_fit.converged,
            'deviance': float(nb_fit.deviance),
            'pearson_chi2': float(nb_fit.pearson_chi2),
            'mean_citations': float(mean_y),
            'median_citations': float(y_count.median()),
            'coefficients': {}
        }
        
        for param in nb_fit.params.index:
            coef = nb_fit.params[param]
            se = nb_fit.bse[param]
            z_val = coef / se if se > 0 else np.nan
            pval = nb_fit.pvalues[param] if param in nb_fit.pvalues else 2 * (1 - stats.norm.cdf(abs(z_val)))
            
            irr = np.exp(coef)
            irr_lower = np.exp(coef - 1.96 * se)
            irr_upper = np.exp(coef + 1.96 * se)
            
            results['part2_count']['coefficients'][param] = {
                'coefficient': float(coef),
                'std_error': float(se),
                'z_value': float(z_val),
                'irr': float(irr),
                'irr_ci_lower': float(irr_lower),
                'irr_ci_upper': float(irr_upper),
                'p_value': float(pval),
                'significant_05': bool(pval < 0.05),
                'significant_01': bool(pval < 0.01),
                'significant_001': bool(pval < 0.001)
            }
        
        log(f"Part 2 converged: {nb_fit.converged}, n={len(y_count)}")
        
    except Exception as e:
        log(f"Error in Part 2 (Negative Binomial): {e}", "ERROR")
        results['part2_count']['error'] = str(e)
    
    return results


# Legacy function name for compatibility
def zinb_regression(df: pd.DataFrame, predictors: List[str]) -> Dict:
    """Wrapper that calls the new hurdle model."""
    return hurdle_model_citations(df, predictors)

# =============================================================================
# VISUALIZATION - Publication Ready
# =============================================================================

# Color-blind friendly palette (Wong, 2011 - Nature Methods)
COLORS = {
    'blue': '#0072B2',      # Blue
    'orange': '#E69F00',    # Orange  
    'green': '#009E73',     # Bluish green
    'vermillion': '#D55E00', # Vermillion (red-orange)
    'purple': '#CC79A7',    # Reddish purple
    'sky': '#56B4E9',       # Sky blue
    'yellow': '#F0E442',    # Yellow
    'black': '#000000'      # Black
}

def set_publication_style():
    """Set matplotlib style for publication-ready figures."""
    plt.rcParams.update({
        'font.family': 'Arial',
        'font.weight': 'bold',
        'font.size': 11,
        'axes.labelsize': 12,
        'axes.labelweight': 'bold',
        'axes.titlesize': 14,
        'axes.titleweight': 'bold',
        'xtick.labelsize': 10,
        'ytick.labelsize': 10,
        'legend.fontsize': 10,
        'figure.dpi': 600,
        'savefig.dpi': 600,
        'savefig.bbox': 'tight',
        'savefig.pad_inches': 0.1,
        'axes.spines.top': False,
        'axes.spines.right': False,
    })

def plot_posters_over_years(df: pd.DataFrame, output_path: Path) -> None:
    """Create bar chart of posters shared per year - publication ready."""
    log("Creating posters-over-years visualization...")
    set_publication_style()
    
    year_counts = df['creation_year'].value_counts().sort_index()
    year_counts = year_counts[(year_counts.index >= 2008) & (year_counts.index <= 2024)]
    
    fig, ax = plt.subplots(figsize=(10, 5))
    bars = ax.bar(year_counts.index, year_counts.values, 
                  color=COLORS['blue'], edgecolor='white', linewidth=0.5)
    
    ax.set_xlabel('Year')
    ax.set_ylabel('Number of Posters')
    # No title - will use figure caption
    
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, p: format(int(x), ',')))
    ax.set_xticks(year_counts.index[::2])  # Every other year for readability
    ax.grid(axis='y', alpha=0.3, linestyle='--')
    
    plt.tight_layout()
    plt.savefig(output_path / 'figure_posters_by_year.png', dpi=600, bbox_inches='tight', 
                facecolor='white', edgecolor='none')
    plt.close()
    log(f"Saved posters-by-year figure (600 DPI)")

def plot_citations_over_years(df_citations: pd.DataFrame, output_path: Path) -> None:
    """Create bar chart of citations per year - publication ready."""
    log("Creating citations-over-years visualization...")
    set_publication_style()
    
    df_citations = df_citations.copy()
    df_citations['citation_year'] = pd.to_numeric(df_citations['citation_year'], errors='coerce')
    year_counts = df_citations['citation_year'].dropna().astype(int).value_counts().sort_index()
    year_counts = year_counts[(year_counts.index >= 2010) & (year_counts.index <= 2025)]
    
    fig, ax = plt.subplots(figsize=(10, 5))
    bars = ax.bar(year_counts.index, year_counts.values, 
                  color=COLORS['vermillion'], edgecolor='white', linewidth=0.5)
    
    ax.set_xlabel('Citation Year')
    ax.set_ylabel('Number of Citations')
    # No title - will use figure caption
    
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, p: format(int(x), ',')))
    ax.grid(axis='y', alpha=0.3, linestyle='--')
    
    plt.tight_layout()
    plt.savefig(output_path / 'figure_citations_by_year.png', dpi=600, bbox_inches='tight',
                facecolor='white', edgecolor='none')
    plt.close()
    log(f"Saved citations-by-year figure (600 DPI)")

def plot_engagement_distribution(df: pd.DataFrame, output_path: Path) -> None:
    """Create separate cumulative distribution plots for each engagement metric."""
    log("Creating engagement distribution visualizations (3 separate figures)...")
    set_publication_style()
    
    metrics_config = [
        ('views', COLORS['blue'], [1, 5, 10], 'Views'),
        ('downloads', COLORS['green'], [1, 5, 10], 'Downloads'),
        ('citations', COLORS['vermillion'], [1, 5, 10], 'Citations')
    ]
    
    for metric, color, points, label in metrics_config:
        if metric not in df.columns:
            continue
        
        fig, ax = plt.subplots(figsize=(8, 6))
        
        x = df[metric].fillna(0).clip(lower=0).to_numpy()
        x_sorted = np.sort(x)[::-1]
        total = x_sorted.sum()
        
        if total <= 0:
            ax.text(0.5, 0.5, f"No {metric} data", ha='center', va='center', transform=ax.transAxes)
            plt.close()
            continue
        
        cum_share = np.cumsum(x_sorted) / total
        top_pct = np.arange(1, len(x_sorted) + 1) / len(x_sorted)
        
        # Main line
        ax.plot(top_pct * 100, cum_share * 100, color=color, linewidth=2.5)
        
        # Annotate key points
        for p in points:
            idx = int(np.ceil(len(x_sorted) * (p / 100))) - 1
            idx = max(0, min(idx, len(x_sorted) - 1))
            share_val = cum_share[idx] * 100
            
            ax.scatter([p], [share_val], color=color, s=80, zorder=5, edgecolor='white', linewidth=1.5)
            
            # Adjust annotation position based on point
            offset_x = 2 if p < 50 else -2
            ha = 'left' if p < 50 else 'right'
            ax.annotate(f"{share_val:.0f}%", 
                       (p + offset_x, share_val), 
                       fontsize=11, fontweight='bold', ha=ha, va='center')
        
        ax.set_xlabel('Top Percentage of Posters (Ranked by ' + label + ')')
        ax.set_ylabel(f'Cumulative Share of Total {label} (%)')
        ax.grid(True, alpha=0.3, linestyle='--')
        ax.set_xlim(0, 100)
        ax.set_ylim(0, 105)
        
        # Add reference line at 50%
        ax.axhline(y=50, color='gray', linestyle=':', alpha=0.5, linewidth=1)
        
        plt.tight_layout()
        plt.savefig(output_path / f'figure_concentration_{metric}.png', dpi=600, 
                    bbox_inches='tight', facecolor='white', edgecolor='none')
        plt.close()
        log(f"Saved concentration figure for {metric} (600 DPI)")

# =============================================================================
# RESULTS EXPORT
# =============================================================================

def export_results(results: Dict, output_path: Path) -> None:
    """Export all results to CSV and JSON files."""
    log("Exporting results...")
    
    # Table 2 - Distribution stats
    if 'table2' in results:
        results['table2'].to_csv(output_path / 'table2_distribution_stats.csv', index=False)
        log("Saved table2_distribution_stats.csv")
    
    # Bivariate results
    for key in ['bivariate_views_spearman', 'bivariate_views_mannwhitney',
                'bivariate_downloads_spearman', 'bivariate_downloads_mannwhitney',
                'bivariate_citations_binary_pointbiserial', 'bivariate_citations_binary_chisquare',
                'bivariate_citations_count_spearman', 'bivariate_citations_count_mannwhitney']:
        if key in results and isinstance(results[key], pd.DataFrame) and len(results[key]) > 0:
            results[key].to_csv(output_path / f'{key}.csv', index=False)
            log(f"Saved {key}.csv")
    
    # Multivariate results as JSON
    for key in ['nb_views', 'nb_downloads', 'zinb_citations']:
        if key in results:
            with open(output_path / f'{key}_results.json', 'w') as f:
                json.dump(results[key], f, indent=2, default=str)
            log(f"Saved {key}_results.json")
    
    # Cleaning stats - convert numpy types to native Python
    if 'citation_cleaning_stats' in results:
        # Use .item() to preserve float precision (not int() which truncates floats)
        clean_stats = {k: v.item() if hasattr(v, 'item') else v 
                       for k, v in results['citation_cleaning_stats'].items()}
        with open(output_path / 'citation_cleaning_stats.json', 'w') as f:
            json.dump(clean_stats, f, indent=2)
        log("Saved citation_cleaning_stats.json")
    
    # Export Excel workbook with formatted tables
    export_excel_tables(results, output_path)


def export_excel_tables(results: Dict, output_path: Path) -> None:
    """
    Export all analysis tables to a formatted Excel workbook.
    Formatting matches the paper tables (Arial 10pt, bold headers, borders).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        log("openpyxl not installed, skipping Excel export", "WARNING")
        return
    
    log("Generating Excel workbook...")
    
    excel_path = output_path / "analysis_tables.xlsx"
    
    # Define consistent styling (Arial 10pt, matching docx)
    header_font = Font(name='Arial', size=10, bold=True)
    cell_font = Font(name='Arial', size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    def style_worksheet(ws, header_row=1):
        """Apply consistent styling to worksheet"""
        for row in ws.iter_rows():
            for cell in row:
                cell.font = cell_font
                cell.border = thin_border
                if cell.row == header_row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                elif cell.column == 1:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb = Workbook()
    
    # =========================================================================
    # TABLE 4: Distribution of Engagement Metrics
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Table 4 - Engagement Dist"
    
    headers = ['Metric', 'N', 'Mean', 'Median', 'Std Dev', '% Zero', 'Top 1% Share', 'Top 10% Share']
    for col, header in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=header)
    
    if 'table2' in results:
        t2 = results['table2']
        for row_idx, (_, row) in enumerate(t2.iterrows(), 2):
            ws1.cell(row=row_idx, column=1, value=row['variable'])
            ws1.cell(row=row_idx, column=2, value=int(row['n']))
            ws1.cell(row=row_idx, column=3, value=round(row['mean'], 1))
            ws1.cell(row=row_idx, column=4, value=int(row['median']))
            ws1.cell(row=row_idx, column=5, value=round(row['std'], 1))
            ws1.cell(row=row_idx, column=6, value=f"{row['pct_zero']:.1f}%")
            ws1.cell(row=row_idx, column=7, value=f"{row['top_1pct_share']:.1f}%")
            ws1.cell(row=row_idx, column=8, value=f"{row['top_10pct_share']:.1f}%")
    
    style_worksheet(ws1)
    
    # =========================================================================
    # TABLE 5: Multivariate Predictors of Views
    # =========================================================================
    ws2 = wb.create_sheet("Table 5 - Views Model")
    
    headers = ['Predictor', 'IRR', '95% CI Lower', '95% CI Upper', 'p-value', 'Sig']
    for col, header in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=header)
    
    key_predictors = [
        ('has_description', 'Has description'),
        ('has_license', 'Has license'),
        ('authors_orcid_percentage', 'ORCID % (per SD)'),
        ('authors_affiliation_percentage', 'Affiliation % (per SD)'),
        ('description_words_count', 'Description words (per SD)'),
        ('keywords_freetext_count', 'Free-text keywords (per SD)'),
        ('has_conference_acronym', 'Has conference acronym'),
        ('has_conference_place', 'Has conference place'),
        ('has_conference_title', 'Has conference title'),
        ('has_conference_session', 'Has conference session'),
        ('has_funding_info', 'Has funding info'),
    ]
    
    if 'nb_views' in results and 'coefficients' in results['nb_views']:
        coeffs = results['nb_views']['coefficients']
        row_idx = 2
        for pred_key, pred_name in key_predictors:
            if pred_key in coeffs:
                c = coeffs[pred_key]
                irr = c['irr']
                ci_lo = c['irr_ci_lower']
                ci_hi = c['irr_ci_upper']
                pval = c['p_value']
                
                if ci_hi > 100:  # Skip unstable estimates
                    continue
                
                sig = '***' if pval < 0.001 else '**' if pval < 0.01 else '*' if pval < 0.05 else ''
                
                ws2.cell(row=row_idx, column=1, value=pred_name)
                ws2.cell(row=row_idx, column=2, value=round(irr, 3))
                ws2.cell(row=row_idx, column=3, value=round(ci_lo, 3))
                ws2.cell(row=row_idx, column=4, value=round(ci_hi, 3))
                ws2.cell(row=row_idx, column=5, value='<0.0001' if pval < 0.0001 else round(pval, 4))
                ws2.cell(row=row_idx, column=6, value=sig)
                row_idx += 1
    
    style_worksheet(ws2)
    
    # =========================================================================
    # TABLE 6: Hurdle Model for Citations
    # =========================================================================
    ws3 = wb.create_sheet("Table 6 - Citations Hurdle")
    
    headers = ['Predictor', 'Part 1: OR', 'Part 1: p', 'Part 1: Sig', 
               'Part 2: IRR', 'Part 2: p', 'Part 2: Sig']
    for col, header in enumerate(headers, 1):
        ws3.cell(row=1, column=col, value=header)
    
    key_predictors_cite = [
        ('has_description', 'Has description'),
        ('has_license', 'Has license'),
        ('keywords_freetext_count', 'Free-text keywords (per SD)'),
        ('references_with_identifiers_count', 'Refs with identifiers (per SD)'),
        ('has_conference_acronym', 'Has conference acronym'),
        ('has_conference_dates', 'Has conference dates'),
        ('has_funding_info', 'Has funding info'),
        ('has_conference_session', 'Has conference session'),
    ]
    
    if 'zinb_citations' in results:
        part1 = results['zinb_citations'].get('part1_logistic', {}).get('coefficients', {})
        part2 = results['zinb_citations'].get('part2_count', {}).get('coefficients', {})
        
        row_idx = 2
        for pred_key, pred_name in key_predictors_cite:
            ws3.cell(row=row_idx, column=1, value=pred_name)
            
            # Part 1
            if pred_key in part1:
                c = part1[pred_key]
                or_val = c['odds_ratio']
                pval = c['p_value']
                sig = '***' if pval < 0.001 else '**' if pval < 0.01 else '*' if pval < 0.05 else ''
                ws3.cell(row=row_idx, column=2, value=round(or_val, 3))
                ws3.cell(row=row_idx, column=3, value='<0.001' if pval < 0.001 else round(pval, 3))
                ws3.cell(row=row_idx, column=4, value=sig)
            else:
                ws3.cell(row=row_idx, column=2, value='—')
                ws3.cell(row=row_idx, column=3, value='—')
            
            # Part 2
            if pred_key in part2:
                c = part2[pred_key]
                irr = c['irr']
                pval = c['p_value']
                se = c.get('std_error', c.get('robust_std_error', 0))
                sig = '***' if pval < 0.001 else '**' if pval < 0.01 else '*' if pval < 0.05 else ''
                if se < 1000:
                    ws3.cell(row=row_idx, column=5, value=round(irr, 3))
                    ws3.cell(row=row_idx, column=6, value='<0.001' if pval < 0.001 else round(pval, 3))
                    ws3.cell(row=row_idx, column=7, value=sig)
                else:
                    ws3.cell(row=row_idx, column=5, value='—')
                    ws3.cell(row=row_idx, column=6, value='—')
            else:
                ws3.cell(row=row_idx, column=5, value='—')
                ws3.cell(row=row_idx, column=6, value='—')
            
            row_idx += 1
    
    style_worksheet(ws3)
    
    # =========================================================================
    # BIVARIATE RESULTS
    # =========================================================================
    ws4 = wb.create_sheet("Bivariate - Spearman")
    
    headers = ['Predictor', 'Views ρ', 'Views p', 'Downloads ρ', 'Downloads p']
    for col, header in enumerate(headers, 1):
        ws4.cell(row=1, column=col, value=header)
    
    if 'bivariate_views_spearman' in results and 'bivariate_downloads_spearman' in results:
        views_df = results['bivariate_views_spearman']
        downloads_df = results['bivariate_downloads_spearman']
        
        row_idx = 2
        for _, row in views_df.iterrows():
            pred = row['predictor']
            downloads_row = downloads_df[downloads_df['predictor'] == pred]
            
            ws4.cell(row=row_idx, column=1, value=pred)
            ws4.cell(row=row_idx, column=2, value=round(row['statistic'], 3))
            ws4.cell(row=row_idx, column=3, value='<0.001' if row['p_value'] < 0.001 else round(row['p_value'], 4))
            
            if len(downloads_row) > 0:
                ws4.cell(row=row_idx, column=4, value=round(downloads_row.iloc[0]['statistic'], 3))
                pval = downloads_row.iloc[0]['p_value']
                ws4.cell(row=row_idx, column=5, value='<0.001' if pval < 0.001 else round(pval, 4))
            
            row_idx += 1
    
    style_worksheet(ws4)
    
    # Save workbook
    wb.save(excel_path)
    log(f"Saved analysis_tables.xlsx with {len(wb.sheetnames)} sheets")


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """Main execution function."""
    log_section("POSTER REUSE ANALYSIS")
    log(f"Output directory: {OUTPUT_PATH}")
    
    # Create output directory
    OUTPUT_PATH.mkdir(parents=True, exist_ok=True)
    
    results = {}
    
    # -------------------------------------------------------------------------
    # PHASE 1: Data Loading and Cleaning
    # -------------------------------------------------------------------------
    log_section("PHASE 1: DATA LOADING AND CLEANING")
    
    # Load data
    df_posters = load_combined_data()
    df_citations = load_citations_ndjson()
    
    # Normalize DOI column for matching
    if 'doi_norm' not in df_posters.columns:
        df_posters['doi_norm'] = df_posters['doi'].str.lower()
    
    # Clean citations
    df_citations_clean, cleaning_stats = clean_citations(df_citations, df_posters)
    results['citation_cleaning_stats'] = cleaning_stats
    
    # Save cleaned citations
    save_clean_citations(df_citations_clean, OUTPUT_PATH)
    
    # Update citation counts
    df_posters = update_citation_counts(df_posters, df_citations_clean)
    
    # Use cleaned citations for analysis
    if 'citations_clean' in df_posters.columns:
        df_posters['citations'] = df_posters['citations_clean']
    
    # -------------------------------------------------------------------------
    # PHASE 2: Descriptive Statistics
    # -------------------------------------------------------------------------
    log_section("PHASE 2: DESCRIPTIVE STATISTICS")
    
    # Generate Table 2
    results['table2'] = generate_table2(df_posters)
    print("\n--- Table 2: Distribution of Engagement Metrics ---")
    print(results['table2'].to_string(index=False))
    
    # -------------------------------------------------------------------------
    # PHASE 3: Visualizations
    # -------------------------------------------------------------------------
    log_section("PHASE 3: VISUALIZATIONS")
    
    plot_posters_over_years(df_posters, OUTPUT_PATH)
    plot_citations_over_years(df_citations_clean, OUTPUT_PATH)
    plot_engagement_distribution(df_posters, OUTPUT_PATH)
    
    # -------------------------------------------------------------------------
    # PHASE 4: Bivariate Analysis - Views & Downloads
    # -------------------------------------------------------------------------
    log_section("PHASE 4: BIVARIATE ANALYSIS - VIEWS & DOWNLOADS")
    
    for outcome in ['views', 'downloads']:
        log(f"\n--- Analyzing {outcome.upper()} ---")
        
        # Spearman for continuous
        spearman = bivariate_continuous_spearman(df_posters, outcome, METADATA_COLS_CONTINUOUS)
        results[f'bivariate_{outcome}_spearman'] = spearman
        if len(spearman) > 0:
            sig = spearman[spearman['significant_05']]
            log(f"Spearman: {len(sig)}/{len(spearman)} predictors significant (p<0.05)")
        
        # Mann-Whitney for boolean
        mannwhitney = bivariate_boolean_mannwhitney(df_posters, outcome, METADATA_COLS_BOOLEAN)
        results[f'bivariate_{outcome}_mannwhitney'] = mannwhitney
        if len(mannwhitney) > 0:
            sig = mannwhitney[mannwhitney['significant_05']]
            log(f"Mann-Whitney: {len(sig)}/{len(mannwhitney)} predictors significant (p<0.05)")
    
    # -------------------------------------------------------------------------
    # PHASE 5: Bivariate Analysis - Citations (Binary)
    # -------------------------------------------------------------------------
    log_section("PHASE 5: BIVARIATE ANALYSIS - CITATIONS (BINARY)")
    
    # Point-biserial for continuous
    pointbiserial = bivariate_binary_pointbiserial(df_posters, METADATA_COLS_CONTINUOUS)
    results['bivariate_citations_binary_pointbiserial'] = pointbiserial
    if len(pointbiserial) > 0:
        sig = pointbiserial[pointbiserial['significant_05']]
        log(f"Point-Biserial: {len(sig)}/{len(pointbiserial)} predictors significant (p<0.05)")
    
    # Chi-square for boolean
    chisquare = bivariate_binary_chisquare(df_posters, METADATA_COLS_BOOLEAN)
    results['bivariate_citations_binary_chisquare'] = chisquare
    if len(chisquare) > 0:
        sig = chisquare[chisquare['significant_05']]
        log(f"Chi-Square: {len(sig)}/{len(chisquare)} predictors significant (p<0.05)")
    
    # -------------------------------------------------------------------------
    # PHASE 6: Bivariate Analysis - Citations (Count)
    # -------------------------------------------------------------------------
    log_section("PHASE 6: BIVARIATE ANALYSIS - CITATIONS (COUNT)")
    
    spearman_cit, mannwhitney_cit = bivariate_citations_count(df_posters)
    results['bivariate_citations_count_spearman'] = spearman_cit
    results['bivariate_citations_count_mannwhitney'] = mannwhitney_cit
    
    if len(spearman_cit) > 0:
        sig = spearman_cit[spearman_cit['significant_05']]
        log(f"Spearman (cited only): {len(sig)}/{len(spearman_cit)} predictors significant")
    
    # -------------------------------------------------------------------------
    # PHASE 7: Multivariate Analysis - Negative Binomial
    # -------------------------------------------------------------------------
    log_section("PHASE 7: MULTIVARIATE ANALYSIS - NEGATIVE BINOMIAL")
    
    all_predictors = METADATA_COLS_CONTINUOUS + METADATA_COLS_BOOLEAN
    
    # Views
    results['nb_views'] = negative_binomial_regression(df_posters, 'views', all_predictors)
    
    # Downloads
    results['nb_downloads'] = negative_binomial_regression(df_posters, 'downloads', all_predictors)
    
    # -------------------------------------------------------------------------
    # PHASE 8: Multivariate Analysis - ZINB for Citations
    # -------------------------------------------------------------------------
    log_section("PHASE 8: MULTIVARIATE ANALYSIS - ZERO-INFLATED NEGATIVE BINOMIAL")
    
    results['zinb_citations'] = zinb_regression(df_posters, all_predictors)
    
    # -------------------------------------------------------------------------
    # PHASE 9: Export Results
    # -------------------------------------------------------------------------
    log_section("PHASE 9: EXPORTING RESULTS")
    
    export_results(results, OUTPUT_PATH)
    
    # Save updated poster table
    df_posters.to_csv(OUTPUT_PATH / 'combined-table-analyzed.csv', index=False)
    log("Saved combined-table-analyzed.csv")
    
    # -------------------------------------------------------------------------
    # Summary
    # -------------------------------------------------------------------------
    log_section("ANALYSIS COMPLETE")
    log(f"Total posters analyzed: {len(df_posters):,}")
    log(f"Total citations (cleaned): {df_posters['citations'].sum():,}")
    log(f"Posters with citations: {(df_posters['citations'] > 0).sum():,}")
    log(f"Results saved to: {OUTPUT_PATH}")
    
    print("\n" + "=" * 80)
    print("Files generated:")
    for f in sorted(OUTPUT_PATH.glob('*')):
        print(f"  - {f.name}")
    print("=" * 80)

if __name__ == "__main__":
    main()

